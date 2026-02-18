from typing import List, Dict
from loguru import logger
import json
import csv
from app.rag.table_processor import TableProcessor
class DocumentLoader:
    """多格式文档加载器"""
    @staticmethod
    def load(file_path:str)->str:
        ext=file_path.lower().split('.')[-1]
        loaders={
            'txt': DocumentLoader._load_text,
            'md': DocumentLoader._load_text,
            'pdf': DocumentLoader.load_pdf,
            'docx': DocumentLoader.load_docx,
            'html': DocumentLoader.load_html,
            'htm': DocumentLoader.load_html,
            'csv': DocumentLoader.load_csv,
            'json': DocumentLoader.load_json,
            'xlsx': DocumentLoader.load_excel,  # 新增
            'xls': DocumentLoader.load_excel,
        }
        loader=loaders.get(ext)
        if loader is None:
            raise ValueError(f"不支持的文件格式：{ext}")
        return loader(file_path)
    @staticmethod
    def _load_text(file_path:str)->str:
        with open(file_path,'r',encoding='utf-8') as f:
            return f.read()
    @staticmethod
    def load_pdf(file_path:str)->str:
        from pypdf import PdfReader
        reader=PdfReader(file_path)
        text_parts=[]
        for i,page in enumerate(reader.pages):
            text=page.extract_text()
            if text:
                text_parts.append(f"[第{i+1}页]\n{text}")
        return "\n\n".join(text_parts)
    @staticmethod
    def load_docx(file_path:str)->str:
        from docx import Document
        doc=Document(file_path)
        text_parts=[]
        for para in doc.paragraphs:
            if para.text.strip():
                text_parts.append(para.text)
        return "\n\n".join(text_parts)
    @staticmethod
    def load_html(file_path:str)->str:
        from bs4 import BeautifulSoup
        with open(file_path,'r',encoding='utf-8') as f:
            soup=BeautifulSoup(f.read(),'lxml')
        for tag in soup(['script','style','nav','footer']):
            tag.decompose()
        return soup.get_text(separator='\n',strip=True)
    @staticmethod
    def load_csv(file_path:str)->str:
        """加载csv文件，使用语义化表格处理"""
        headers=[]
        rows=[]
        with open(file_path,'r',encoding='utf-8') as f:
            reader=csv.reader(f)
            headers=next(reader,[])
            rows=[row for row in reader]
            table_name=file_path.split('/')[-1].split('\\')[-1]
            return TableProcessor.to_semantic_text(headers,rows,table_name)
        
    @staticmethod  
    def load_json(file_path:str)->str:
        """加载json"""
        with open(file_path,'r',encoding='utf-8') as f:
            data=json.load(f)
        if isinstance(data,list):
            text_parts=[f"json数组，共{len(data)}个元素"]
            for i,item in enumerate(data):
                text_parts.append(f"[{i+1}] {json.dumps(item,ensure_ascii=False)}")
            return '\n'.join(text_parts)
        elif isinstance(data,dict):
            return f"json对象：\n{json.dumps(data,ensure_ascii=False,indent=2)}"
        else:
            return str(data)
    @staticmethod
    def load_excel(file_path: str) -> str:
        """加载 Excel 文件"""
        from openpyxl import load_workbook
        wb=load_workbook(file_path,data_only=True)
        text_parts=[]
        for sheet_name in wb.sheetnames:
            sheet=wb[sheet_name]

            rows=[]
            for row in sheet.iter_rows(values_only=True):
                row_data=[str(cell) if cell is not None else '' for cell in row]
                if any(cell.strip() for cell in row_data):
                    rows.append(row_data)
            if rows:
                headers=rows[0]
                data_rows=rows[1:]
                sheet_text=TableProcessor.to_semantic_text(headers,data_rows,table_name=f"{sheet_name}")
                text_parts.append(sheet_text)
        return "\n\n".join(text_parts)
